from .models import *
import pandas as pd
from django.http import HttpResponse
from openpyxl import load_workbook
from .views_salary import salary_logic
from openpyxl.styles import Font,Border, Side,Alignment
from .extra import showing_adminpageindex,int_to_month_string_uz
from openpyxl.cell import MergedCell

def styling_excell(data, response, last_line=False, title=""):
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        # Write data to Excel, starting from row 2 (since row 1 is reserved for title)
        data.to_excel(writer, sheet_name='Oylik', index=False, startrow=1)
        workbook = writer.book
        worksheet = writer.sheets['Oylik']

        # Formatting elements
        bold_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Insert the title in the first row
        num_columns = data.shape[1]  # Get number of columns
        last_column_letter = chr(65 + num_columns - 1)  # Convert column number to letter
        merge_range = f"A1:{last_column_letter}1"  # Merge from A1 to last column in first row
        worksheet.merge_cells(merge_range)
        title_cell = worksheet["A1"]
        title_cell.value = title
        title_cell.font = Font(bold=True, size=8)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Apply borders and center alignment to all data cells
        for row in worksheet.iter_rows(min_row=2):  # Start from row 2 to skip title row
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Make the header row (row 2) bold
        for cell in worksheet[3]:  
            cell.font = bold_font

        # Make the last row bold if required
        if last_line:
            last_row = worksheet.max_row
            for cell in worksheet[last_row]:
                cell.font = bold_font

        # Adjust column widths (excluding the first row with merged title)
        for col in worksheet.iter_cols(min_row=2):  # Start from row 2 to avoid merged cell issue
            max_length = 0
            col_letter = col[0].column_letter  # Get column letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2  # Add some padding
            worksheet.column_dimensions[col_letter].width = adjusted_width

    return response

def main_to_excell(request,month,year,typeid):
    company = request.user.company
    title= f"1-{int_to_month_string_uz(month=month)}дан 1-{int_to_month_string_uz(month=month+1)}гача ИШЧИЛАНИНГ БАЖАРГАН ИШЛАРИ, маҳсулот - {Progresstype.objects.get(id = typeid).type}"
    
    dt = showing_adminpageindex(company,month,year,typeid)
    main = pd.DataFrame(dt)
   

    # Append the new row
    

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="hisobchi.xlsx"'
    return styling_excell(main,response,last_line=True,title = title)


def salaries_to_excell(request,pk):
    d = DateforProgress.objects.filter(company =request.user.company )[0]
    month = d.date.month
    title= f"{int_to_month_string_uz(month)} ойида ишлаган ишчиларнинг ҳисобланган ойлиги"
    data = salary_logic(cp=request.user.company,pk=pk)
    dt = []
    k=1
    for i in data:
        
        info = {"№ ":k,"Ф.И.Ш.":i["worker"],
                }
        for ex in i["calculating_expanses"]:
            info.update({f"{ex["expanse_name"]}":ex["expanse_sum"]})
        info.update({
            "Қилинган иш":i["done_money"],
            "Аванс":i["avans"],
            "Премя":i["premya"],
            "Қўлга тегиши":i["full_money"],
            "имзо":"            "
        })
        
        
        dt.append(info)
        k+=1

    main = pd.DataFrame(dt)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="progress_data.xlsx"'
    return styling_excell(main,response,title=title)


def styling_excell2(worksheet, data, title=""):
    # ✅ Ensure DataFrame is not empty before writing
    if data.empty:
        data = pd.DataFrame({"No Data": ["No records found"]})

    # ✅ Ensure at least one visible sheet
    worksheet.sheet_state = 'visible'

    # Formatting elements
    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    num_columns = data.shape[1]
    last_column_letter = chr(65 + num_columns - 1)
    merge_range = f"A1:{last_column_letter}1"
    worksheet.merge_cells(merge_range)

    title_cell = worksheet["A1"]
    title_cell.value = title
    title_cell.font = Font(bold=True, size=8)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # ✅ Apply borders & center alignment
    for row in worksheet.iter_rows(min_row=2):  
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
    for cell in worksheet[3]:  
            cell.font = bold_font

        # Make the last row bold if required
    last_line = True
    if last_line:
        last_row = worksheet.max_row
        for cell in worksheet[last_row]:
            cell.font = bold_font
    # ✅ Auto-adjust column width
    for col in worksheet.iter_cols(min_row=2):
        col_letter = col[0].column_letter
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        worksheet.column_dimensions[col_letter].width = max_length + 2


def admindate_main_to_excel(request, pk):
    date = DateforProgress.objects.get(id=pk)
    month = date.date.month
    year = date.date.year
    company = request.user.company
    type_ids = Progresstype.objects.filter(date=date)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="hisobchi.xlsx"'

    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        for typeid in type_ids:
            title = f"1-{int_to_month_string_uz(month)}дан 1-{int_to_month_string_uz(month+1)}гача ИШЧИЛАНИНГ БАЖАРГАН ИШЛАРИ, маҳсулот - {typeid.type}"
            dt = showing_adminpageindex(company, month, year, typeid.id)
            df = pd.DataFrame(dt)

            sheet_name = f"{typeid.type}"
            df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)

            # ✅ Apply styling after writing the data
            worksheet = writer.sheets[sheet_name]
            styling_excell2(worksheet, df, title)

    return response
